
import json
import openpyxl
import gspread
from googleapiclient import discovery
from google.oauth2 import service_account

import google.auth.exceptions
import requests.exceptions
import time
import threading
import queue

import os

class Filhandterer:

    def __init__(self, ordark_navn, excel_specs, googleark_specs):

        self._ordark_navn = ordark_navn
        self.les_ordark = None
        self.les_ordark2 = None
        self.skriv_ordark = []
   
        excel_filnavn, excel_les, excel_skriv = excel_specs.values()
        googleark_id, googleark_les, googleark_skriv = googleark_specs.values()

        # excel
        self._excel_filnavn = "mine-gloser/" + excel_filnavn
        self._excel_fil = None

        # googleark
        self._google_service = None
        self._googleark_id = googleark_id
        self._googleark_fil = None

        if excel_les:
            if googleark_les:
                print("FEIL: Både excel og googleark var valgt til å leses fra. Avbryter.")
                exit(1)
            self.les_ordark = self.les_excel
        elif googleark_les:
            self.les_ordark = self.les_googleark
        else:
            print("FEIL: Verken excel eller googleark var valgt til å leses fra. Avbryter.")
            exit(1)

        if excel_skriv:
            self.skriv_ordark.append(self.skriv_excel)
            if (self.les_excel != self.les_ordark):
                self.les_ordark2 = self.les_excel

        if googleark_skriv:
            self.skriv_ordark.append(self.skriv_googleark)
            if (self.les_googleark != self.les_ordark):
                self.les_ordark2 = self.les_googleark

    def initialiser_excel(self, hensikt):
        if self._excel_fil != None:
            return True
        self._excel_fil = self.brukerfeil_vent(PermissionError, hensikt, openpyxl.load_workbook, self._excel_filnavn)
        return True

    def initialiser_googleark(self, hensikt, kan_skippe_feil=False):
        if self._googleark_fil != None:
            return True
        scope = ['https://www.googleapis.com/auth/drive']

        credentials = service_account.Credentials.from_service_account_info({
                "type": "service_account",
                "project_id": os.environ.get("GOOGLE_PROJECT_ID"),
                "private_key_id": os.environ.get("GOOGLE_PRIVATE_KEY_ID"),
                "private_key": os.environ.get("GOOGLE_PRIVATE_KEY").replace("\\n", "\n"),
                "client_email": os.environ.get("GOOGLE_CLIENT_EMAIL"),
                "client_id": os.environ.get("GOOGLE_CLIENT_ID"),
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
                "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/test-91%40testnavn.iam.gserviceaccount.com"
        }
            , scopes=scope)
        klient = gspread.authorize(credentials)

        self._google_service = discovery.build('sheets', 'v4', credentials=credentials)
        
        returverdi = self.brukerfeil_vent(google.auth.exceptions.TransportError, hensikt, klient.open_by_key, self._googleark_id, kan_skippe=kan_skippe_feil)
        if returverdi == "hoppet over":
            return False
    
        self._googleark_fil = returverdi
        return True

    @staticmethod
    def les_json(filnavn):
        streng = ""
        with open(filnavn, encoding='utf-8') as fil:
            for linje in fil:
                streng += linje.strip()

        return json.loads(streng)
    
    def les_excel(self):

        if not self.initialiser_excel("Innlesing"):
            return

        arknavn = self._ordark_navn
        if not self.excel_har_ark(arknavn):
            raise ValueError(f'Ordliste-ark "{arknavn}" finnes ikke.')

        ark = self._excel_fil[arknavn]
        
        data = []
        for rad in range(1, ark.max_row + 1):
            data.append([])
            for kol in range(1, ark.max_column + 1):
                verdi = ark.cell(row = rad, column = kol).value
                if verdi == None:
                    data[rad-1].append("")
                    continue
                data[rad-1].append(verdi)

        print(f"[+] Ordliste innhentet fra Excel-fil.")
        return data
    
    def skriv_excel(self, oppdateringer):

        if not self.initialiser_excel("Oppdatering"):
            return
        
        for arknavn, (fanefarge,data) in oppdateringer.items():

            if self.excel_har_ark(arknavn):
                ark = self._excel_fil[arknavn]
            else:
                ark = self._excel_fil.create_sheet(arknavn) 
            
            ark.sheet_properties.tabColor = fanefarge

            for rad_indeks,rad in enumerate(data):
                for kol_indeks,verdi in enumerate(rad):
                    ark.cell(row=rad_indeks+1, column=kol_indeks+1,value=verdi)
                    
            for rad_indeks in range(len(data)+1, ark.max_row + 1):
                for kol_indeks in range(1, ark.max_column + 1):
                    ark.cell(row = rad_indeks, column = kol_indeks).value = None

        self.brukerfeil_vent(PermissionError, "Oppdatering", self._excel_fil.save, self._excel_filnavn)
        print(f"[+] Excel-fil oppdatert.")

    def excel_har_ark(self, arknavn):
        return arknavn in self._excel_fil.sheetnames

    def les_googleark(self):

        if not self.initialiser_googleark("Innlesing"):
            return

        arknavn = self._ordark_navn
        if not self.googleark_har_ark(arknavn):
            raise ValueError(f'Ordliste-ark "{arknavn}" finnes ikke.')

        ark = self.brukerfeil_vent(requests.exceptions.ConnectionError, "lese fra", self._googleark_fil.worksheet, arknavn)
        print(f"[+] Ordliste innhentet fra Google regneark.")
        return ark.get_values()[:]

    def skriv_googleark(self, data):

        if not self.initialiser_googleark("Oppdatering", kan_skippe_feil=True):
            return

        def indre_funk(): 
            verdier = []
            formatteringer = []

            for arknavn,(fanefarge,rader) in data.items():

                if self.googleark_har_ark(arknavn):
                    ark = self._googleark_fil.worksheet(arknavn)
                else:
                    ark = self._googleark_fil.add_worksheet(title=arknavn, rows="100", cols="10")

                formatteringer.append({
                'updateSheetProperties': {
                    'properties': {
                        'sheetId': ark.id,
                        'tabColor': self.hex_til_tab_farge(fanefarge)
                    },
                    'fields': 'tabColor'
                }})

                if (antall_rader := len(rader))>ark.row_count:
                    formatteringer.append({
                    "appendDimension": {
                        "sheetId": ark.id,
                        "dimension": "ROWS",
                        "length": antall_rader-ark.row_count
                    }})

                if antall_rader < ark.row_count:
                    formatteringer.append({
                    "deleteDimension": {
                        "range": {
                            "sheetId": ark.id,
                            "dimension": "ROWS",
                            "startIndex": antall_rader,
                            "endIndex": ark.row_count
                        }
                    }}),

                if (antall_kolonner := len(rader[0]))>ark.col_count:
                    formatteringer.append({
                    "appendDimension": {
                        "sheetId": ark.id,
                        "dimension": "COLUMNS",
                        "length": antall_kolonner-ark.col_count
                    }})

                verdier.append({"range": f"{arknavn}!A1","values": rader})

            if len(formatteringer)>0:
                self._google_service.spreadsheets().batchUpdate(spreadsheetId=self._googleark_id, body={"requests": formatteringer}).execute()
            if len(verdier)>0:
                self._google_service.spreadsheets().values().batchUpdate(spreadsheetId=self._googleark_id, body={"valueInputOption": "RAW","data": verdier}).execute()

        returverdi = self.brukerfeil_vent(requests.exceptions.ConnectionError, "oppdatere", indre_funk, kan_skippe=True)
        if not returverdi == "hoppet over":
            print(f"[+] Google regneark oppdatert.")

    def googleark_har_ark(self, arknavn):
        for sheet in self._googleark_fil.worksheets():
            if sheet.title == arknavn:
                return True
        return False

    def brukerfeil_vent(self, error, hensikt, kommando, *args, kan_skippe=False):
        
        if error == PermissionError:
            eventuelt_varsel = f'VENTER: {hensikt} av Excel-fil er hindret fordi "{args[0]}" er åpen. Venter til filen lukkes.\n'

        elif error in (google.auth.exceptions.TransportError, requests.exceptions.ConnectionError):
            eventuelt_varsel = f'VENTER: {hensikt} av Google regneark er hindret fordi forbindelse til internett mangler. Venter til forbindelsen opprettes.\n'

        try:
            return kommando(*args)
        except error:
            print(eventuelt_varsel)

        if kan_skippe:
            print('Eventuelt kan du hoppe over dette steget ved å skrive "hopp over".')

            def vent_input(channel):
                while True:
                    svar = input("Svar: ")
                    if svar == "hopp over":
                        print("Hopper over steg.")
                        break
                    else:
                        print('Svaret lot seg ikke forstå. Skriv "hopp over" hvis du vil hoppe over steget.')
                channel.put(svar)

            channel = queue.Queue()
            thread = threading.Thread(target=vent_input, args=(channel,))
            thread.daemon = True
            thread.start()
        
        while True:
            try:
                returverdi = kommando(*args)
                if kan_skippe:
                    print()
                return returverdi
            except error:
                if not kan_skippe:
                    time.sleep(1)
                    continue

                try:
                    channel.get(True,1)
                    return "hoppet over"
                except queue.Empty:
                    pass
             
    def hex_til_tab_farge(self, hex_verdi):
        lyshetsgrad = 256 # 256 beholder RGB verdi

        tab_farge = {
            "red": int(hex_verdi[0:2], 16)/lyshetsgrad,  # Replace with the desired RGB values
            "green": int(hex_verdi[2:4], 16)/lyshetsgrad,
            "blue": int(hex_verdi[4:6], 16)/lyshetsgrad,
        }
        return tab_farge